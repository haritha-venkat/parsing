"""
capex_summary.py
─────────────────
Extracts structured capital expenditure data from 5 mining annual reports
using PyMuPDF and outputs a formatted Excel + CSV summary table.

Output columns:
    Company | Year | Total Capex ($M) | Sustaining ($M) |
    Growth ($M) | Development ($M) | Notes | Source Pages
"""

import io
import re
import sys
from pathlib import Path

import fitz  # PyMuPDF
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── PDF targets ────────────────────────────────────────────────────────────────
DATA_DIR = Path("C:/Users/haritha.shree/Downloads/rag_project/rag_project/data")

TARGETS = {
    "Antofagasta": DATA_DIR / "antofagasta-2022-ara.pdf",
    "Barrick Gold": DATA_DIR / "Barrick_Annual_Report_2022.pdf",
    "Evolution Mining": DATA_DIR / "Evolution Mining -Annual-Report-2022.pdf",
    "Newcrest Mining": DATA_DIR / "New Crest - 221004_Annual Report 2022_0.pdf",
    "First Majestic": DATA_DIR / "2022-annual-report - firstmajestic.pdf",
}

# ── Helpers ────────────────────────────────────────────────────────────────────


def get_full_text(pdf_path: Path) -> str:
    """Return all text from a PDF as a single string."""
    doc = fitz.open(str(pdf_path))
    pages_text = []
    for i in range(len(doc)):
        pages_text.append(f"[PAGE {i+1}]\n{doc[i].get_text('text')}")
    doc.close()
    return "\n".join(pages_text)


def find_amount(text: str, patterns: list[str]) -> tuple[float | None, str, str]:
    """
    Try patterns in order. Returns (amount_millions, raw_match, page_ref).
    Normalises billion → million.
    """
    for pat in patterns:
        m = re.search(pat, text, re.IGNORECASE | re.DOTALL)
        if m:
            groups = m.groups()
            # We always store value in group 1, unit in group 2
            val_str, unit = groups[0], groups[1]
            val = float(val_str.replace(",", ""))
            if unit.lower() in ("billion", "b"):
                val *= 1000
            # Find page ref
            page_ref = ""
            page_m = re.search(
                r"\[PAGE (\d+)\]", text[max(0, m.start() - 200) : m.start()]
            )
            if page_m:
                page_ref = f"pg{page_m.group(1)}"
            return round(val, 1), m.group(0)[:120].replace("\n", " "), page_ref
    return None, "", ""


# ── Per-company extraction configs ─────────────────────────────────────────────
# Each entry: list of regex patterns with exactly 2 groups: (value, unit)

COMPANY_PATTERNS = {
    "Antofagasta": {
        # Group total on page 101: "to $1,879.2 million in the current year"
        "total": [
            # Group total: "Capital expenditure in 2022 was $1,879.2 million compared with"
            r"[Cc]apital\s+expenditure\s+in\s+2022\s+was\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+compared",
            r"from\s+\$[\d,\.]+\s*million\s+in\s+2021\s+to\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        # Group sustaining: sum of mines — best found at group level
        "sustaining": [
            r"([\d,\.]+)\s*(million|billion|M\b|B\b)\s+on\s+sustaining\s+capital\s+expenditure",
            r"sustaining\s+capital\s+expenditure\s+(?:and|of|was|were)?\s*\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        "growth": [
            r"([\d,\.]+)\s*(million|billion|M\b|B\b)\s+on\s+growth\s+project",
            r"growth\s+project[s]?\s+([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        "development": [
            r"([\d,\.]+)\s*(million|billion|M\b|B\b)\s+on\s+mine\s+development",
            r"mine\s+development[,\s]+([\d,\.]+)\s*(million|billion|M\b|B\b)",
            r"\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+on\s+mine\s+development",
        ],
    },
    "Barrick Gold": {
        # Page 110: "Attributable capital expenditures for 2022 of $2,417 million"
        "total": [
            r"capital\s+expenditures\s+for\s+2022\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
            r"total\s+attributable\s+(?:gold\s+&?\s+copper\s+)?capital\s+expenditure[s]?\D{0,20}([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        # Page 110: "minesite sustaining capital expenditures of $1,678 million"
        "sustaining": [
            r"(?:attributable\s+)?minesite\s+sustaining\s+capital\s+expenditure[s]?[^.]{0,30}of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
            r"minesite\s+sustaining\D{0,40}([\d,\.]+)\s*(million|billion|M\b|B\b)\s+was\s+higher",
        ],
        "growth": [
            r"expenditures\d?\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+was\s+higher\s+than\s+the\s+guidance\s+range\s+of\s+\$550",
            r"\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+was\s+higher\s+than\s+the\s+guidance\s+range\s+of\s+\$550",
            r"project\s+capital\s+expenditure[s]?[^.]{0,30}([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        "development": [
            r"underground\s+development\D{0,40}([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
    },
    "Evolution Mining": {
        # FY22 (year ended 30 June 2022), figures in AUD millions
        # Page 78: "Capital investment for the year totalled $606.4 million"
        "total": [
            r"[Cc]apital\s+investment\s+for\s+the\s+year\s+totall?ed?\s+\$?([\d,\.]+)\s*(million|billion|M\b|B\b)",
            r"[Tt]otal\s+capital\s+investment\s+was\s+\$?([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        # Page 75/78: "sustaining capital investment... of $147.1 million"
        "sustaining": [
            r"([\d,\.]+)\s*(million|billion|M\b|B\b)\s+(?:\(30\s+June[^)]+\)\s+)?of\s+sustaining\s+capital",
            r"sustaining\s+capital\s+investment\s+\(30\s+June\s+2021[^)]+\)\s+and\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
            r"included\s+\$?([\d,\.]+)\s*(million|billion|M\b|B\b)\s+(?:\([^)]+\)\s+)?of\s+sustaining\s+capital",
        ],
        # Page 75/78: "major capital investment... $459.3 million"
        "growth": [
            r"([\d,\.]+)\s*(million|billion|M\b|B\b)\s+(?:\([^)]+\)\s+)?of\s+major\s+capital\s+investment",
            r"major\s+capital\s+investment[^.]{0,40}\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        "development": [
            r"mine\s+development\s+of\s+\$?([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
    },
    "Newcrest Mining": {
        # FY22 (year ended 30 June 2022)
        # Page 63: "Capital expenditure of $1,417 million"
        "total": [
            r"[Cc]apital\s+expenditure\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+in\s+the\s+current\s+period",
            r"[Cc]apital\s+expenditure\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        # Page 63: "Sustaining capital expenditure of $431 million"
        "sustaining": [
            r"[Ss]ustaining\s+capital\s+expenditure\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        # Page 63: "non-sustaining, capital expenditure of $773 million"
        "growth": [
            r"non.sustaining[^.]{0,40}capital\s+expenditure\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
            r"[Mm]ajor\s+project[^.]{0,40}capital\s+expenditure\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        "development": [
            r"[Pp]roduction\s+stripping\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
    },
    "First Majestic": {
        # 2022: "$59.7 million spent on purchase of PP&E"
        # Investing total: $213.8 million
        "total": [
            r"[Cc]ash\s+used\s+in\s+investing\s+activities\s+of\s+\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
            r"\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+spent\s+on\s+purchase\s+of\s+property",
        ],
        "sustaining": [
            r"\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+for\s+sustaining\s+activit",
            r"sustaining\s+activit[^.]{0,40}\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        "growth": [
            r"\$([\d,\.]+)\s*(million|billion|M\b|B\b)\s+for\s+expansionary\s+project",
            r"expansionary\s+project[s]?[^.]{0,40}\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
        "development": [
            r"mine\s+development[^.]{0,40}\$([\d,\.]+)\s*(million|billion|M\b|B\b)",
        ],
    },
}


# ── Main extraction ────────────────────────────────────────────────────────────


def extract_company_capex(company: str, pdf_path: Path) -> dict:
    row = {
        "Company": company,
        "Year": 2022,
        "Total Capex ($M)": None,
        "Sustaining ($M)": None,
        "Growth ($M)": None,
        "Development ($M)": None,
        "Currency": "USD",
        "Notes": "",
        "Source Pages": "",
    }

    if not pdf_path.exists():
        row["Notes"] = "File not found"
        return row

    text = get_full_text(pdf_path)
    patterns = COMPANY_PATTERNS[company]
    page_refs = []

    for category in ["total", "sustaining", "growth", "development"]:
        val, raw, pg = find_amount(text, patterns.get(category, []))
        col_map = {
            "total": "Total Capex ($M)",
            "sustaining": "Sustaining ($M)",
            "growth": "Growth ($M)",
            "development": "Development ($M)",
        }
        row[col_map[category]] = val
        if pg:
            page_refs.append(pg)

    # Evolution Mining reports in AUD
    if company == "Evolution Mining":
        row["Currency"] = "AUD"
        row["Notes"] = "FY22 (year ended 30 Jun 2022). Figures in AUD millions."
    elif company == "Newcrest Mining":
        row["Notes"] = (
            "FY22 (year ended 30 Jun 2022). Development = Production Stripping."
        )
    elif company == "First Majestic":
        row["Notes"] = (
            "Total = Investing activities outflow. Sustaining/Growth from 2023 plan."
        )
    elif company == "Barrick Gold":
        row["Notes"] = "Attributable basis (incl. JV share). Growth = Project capex."
    elif company == "Antofagasta":
        row["Notes"] = "Group consolidated. Dev = Mine development (Los Pelambres)."

    row["Source Pages"] = ", ".join(sorted(set(page_refs)))
    return row


def main():
    print("Extracting structured capex summary from 5 mining annual reports...\n")

    rows = []
    for company, pdf_path in TARGETS.items():
        print(f"  Processing: {company}")
        row = extract_company_capex(company, pdf_path)
        rows.append(row)
        print(
            f"    Total={row['Total Capex ($M)']}M  "
            f"Sustaining={row['Sustaining ($M)']}M  "
            f"Growth={row['Growth ($M)']}M  "
            f"Dev={row['Development ($M)']}M  "
            f"({row['Currency']})"
        )

    df = pd.DataFrame(
        rows,
        columns=[
            "Company",
            "Year",
            "Currency",
            "Total Capex ($M)",
            "Sustaining ($M)",
            "Growth ($M)",
            "Development ($M)",
            "Notes",
            "Source Pages",
        ],
    )

    # ── Save CSV ───────────────────────────────────────────────────────────────
    csv_path = Path("capex_structured_summary.csv")
    df.to_csv(csv_path, index=False, encoding="utf-8")
    print(f"\nCSV saved  -> {csv_path}")

    # ── Save Excel ─────────────────────────────────────────────────────────────
    excel_path = Path("capex_structured_summary.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Capex Summary")
        ws = writer.sheets["Capex Summary"]

        # Column widths
        for col, width in zip(
            "ABCDEFGHI", [22, 6, 8, 18, 16, 14, 18, 52, 16], strict=True
        ):
            ws.column_dimensions[col].width = width

        # Header row styling
        hdr_fill = PatternFill("solid", fgColor="1F4E79")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = ctr
            cell.border = thin

        # Data rows
        alt_fill = PatternFill("solid", fgColor="DCE6F1")
        wht_fill = PatternFill("solid", fgColor="FFFFFF")
        num_bold = Font(bold=True, size=10)
        reg_font = Font(size=10)

        for r_idx, ws_row in enumerate(ws.iter_rows(min_row=2), start=1):
            fill = alt_fill if r_idx % 2 == 0 else wht_fill
            for cell in ws_row:
                cell.fill = fill
                cell.border = thin
                is_num_col = cell.column in (4, 5, 6, 7)  # D E F G = capex cols
                cell.font = num_bold if is_num_col else reg_font
                cell.alignment = Alignment(
                    horizontal="center" if cell.column <= 7 else "left",
                    vertical="center",
                    wrap_text=True,
                )

        # Freeze header
        ws.freeze_panes = "A2"

    print(f"Excel saved -> {excel_path}")

    # ── Console table ──────────────────────────────────────────────────────────
    print("\n" + "=" * 78)
    print("  MINING COMPANY CAPITAL EXPENDITURE SUMMARY — 2022")
    print("=" * 78)
    display = df[
        [
            "Company",
            "Year",
            "Currency",
            "Total Capex ($M)",
            "Sustaining ($M)",
            "Growth ($M)",
            "Development ($M)",
        ]
    ].copy()
    print(display.to_string(index=False))
    print("=" * 78)
    print("\nNotes:")
    for _, r in df.iterrows():
        if r["Notes"]:
            print(f"  {r['Company']}: {r['Notes']}")


if __name__ == "__main__":
    main()
