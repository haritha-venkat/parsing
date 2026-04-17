"""
capex_table_extractor.py
─────────────────────────
For each mining PDF:
  1. Try to find actual CAPEX TABLES using PyMuPDF find_tables()
  2. If no table found → fallback to text paragraphs containing $ amounts

Saves everything to capex_tables.xlsx (one sheet per company).
"""

import io
import re
import sys
from pathlib import Path

import fitz  # PyMuPDF
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ── Config ─────────────────────────────────────────────────────────────────────
DATA_DIR = Path("C:/Users/haritha.shree/Downloads/rag_project/rag_project/data")

TARGETS = {
    "Antofagasta": DATA_DIR / "antofagasta-2022-ara.pdf",
    "Barrick Gold": DATA_DIR / "Barrick_Annual_Report_2022.pdf",
    "Evolution Mining": DATA_DIR / "Evolution Mining -Annual-Report-2022.pdf",
    "Newcrest Mining": DATA_DIR / "New Crest - 221004_Annual Report 2022_0.pdf",
    "First Majestic": DATA_DIR / "2022-annual-report - firstmajestic.pdf",
}

CAPEX_KW = re.compile(
    r"capital expenditure|capex|sustaining capital|project capital"
    r"|development capital|growth capital|minesite sustaining|major project",
    re.IGNORECASE,
)

AMOUNT_RE = re.compile(
    r"\$[\d,\.]+\s*(?:million|billion|M\b|B\b)|[\d,\.]+\s*(?:million|billion)",
    re.IGNORECASE,
)


# ── Helpers ────────────────────────────────────────────────────────────────────


def clean_cell(val) -> str:
    if val is None:
        return ""
    s = re.sub(r"\n+", " | ", str(val).strip())
    s = re.sub(r"\s{2,}", " ", s)
    # Remove illegal Excel characters (control chars except tab/newline)
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)
    return s[:32767]  # Excel cell limit


def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.map(clean_cell)
    df = df.replace("", pd.NA).dropna(how="all").dropna(axis=1, how="all").fillna("")
    df.columns = [
        f"Col{i+1}" if str(c).startswith("Col") else clean_cell(c)
        for i, c in enumerate(df.columns)
    ]
    return df


CAPEX_ROW_KW = re.compile(
    r"capital expenditure|capex|sustaining capital|project capital"
    r"|development capital|growth capital|minesite sustaining"
    r"|total capital|major project|mine development",
    re.IGNORECASE,
)


def is_capex_table(df: pd.DataFrame) -> bool:
    """
    True ONLY if this table is DEDICATED to capital expenditure.
    Requires at least 2 rows in the first column to match capex keywords.
    This filters out large financial summary tables that just have
    one capex line among many unrelated rows.
    """
    first_col_values = [str(v) for v in df.iloc[:, 0]]
    capex_row_count = sum(1 for v in first_col_values if CAPEX_ROW_KW.search(v))
    return capex_row_count >= 2


# ── Step 1: Try table extraction ───────────────────────────────────────────────


def try_extract_tables(pdf_path: Path) -> list[dict]:
    """Try PyMuPDF find_tables() and return capex tables found."""
    results = []
    doc = fitz.open(str(pdf_path))
    for pg_idx in range(len(doc)):
        page = doc[pg_idx]
        if not CAPEX_KW.search(page.get_text("text")):
            continue
        try:
            tables = page.find_tables()
        except Exception:
            continue
        for t_idx, tbl in enumerate(tables.tables):
            try:
                df = clean_df(tbl.to_pandas())
            except Exception:
                continue
            if df.empty or not is_capex_table(df):
                continue
            results.append({"page": pg_idx + 1, "table_num": t_idx + 1, "df": df})
    doc.close()
    return results


# ── Step 2: Fallback — text paragraphs with $ amounts ─────────────────────────


def fallback_text_extract(pdf_path: Path) -> list[dict]:
    """Extract text lines that contain capex keywords AND dollar amounts."""
    results = []
    doc = fitz.open(str(pdf_path))
    seen = set()

    for pg_idx in range(len(doc)):
        text = doc[pg_idx].get_text("text")
        if not CAPEX_KW.search(text):
            continue

        # Split into paragraphs (double newline or single newline blocks)
        paragraphs = re.split(r"\n{2,}", text)
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            if CAPEX_KW.search(para) and AMOUNT_RE.search(para):
                # Collect only lines with $ amounts
                relevant_lines = []
                for line in para.splitlines():
                    line = line.strip()
                    if AMOUNT_RE.search(line):
                        relevant_lines.append(line)
                if relevant_lines:
                    key = " ".join(relevant_lines)[:80]
                    if key in seen:
                        continue
                    seen.add(key)
                    results.append(
                        {
                            "page": pg_idx + 1,
                            "text": "\n".join(relevant_lines),
                        }
                    )

    doc.close()
    return results


# ── Excel writer ───────────────────────────────────────────────────────────────


def write_excel(all_data: dict[str, list[dict]], out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # Enhanced color scheme
    hdr_fill = PatternFill("solid", fgColor="2E75B6")  # Professional blue
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color="2E75B6", size=12)
    fb_font = Font(bold=True, color="C65911", size=12)  # Orange for fallback
    alt_fill = PatternFill("solid", fgColor="F2F2F2")  # Light gray
    wht_fill = PatternFill("solid", fgColor="FFFFFF")
    data_font = Font(size=10, name="Calibri")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    thick_border = Border(
        left=Side(style="medium", color="2E75B6"),
        right=Side(style="medium", color="2E75B6"),
        top=Side(style="medium", color="2E75B6"),
        bottom=Side(style="medium", color="2E75B6"),
    )

    for company, entries in all_data.items():
        ws = wb.create_sheet(title=company[:31])
        cur = 1

        if not entries:
            ws.cell(cur, 1, f"No capex data found for {company}").font = Font(
                italic=True, size=10
            )
            continue

        # Company header
        company_cell = ws.cell(cur, 1, f"Capital Expenditure Data - {company}")
        company_cell.font = Font(bold=True, size=14, color="2E75B6")
        company_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(
            start_row=cur, start_column=1, end_row=cur, end_column=5
        )  # Merge across columns
        cur += 2

        for entry in entries:
            df = entry["df"]
            page = entry["page"]
            is_fb = entry.get("is_fallback", False)
            label = "TEXT CONTEXT (no table)" if is_fb else "TABLE"

            # Title row
            title = f"Page {page} - {label}"
            cell = ws.cell(cur, 1, title)
            cell.font = fb_font if is_fb else title_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thick_border
            cur += 1

            # Header
            for c_idx, col in enumerate(df.columns, 1):
                hc = ws.cell(cur, c_idx, col)
                hc.fill = hdr_fill
                hc.font = hdr_font
                hc.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                hc.border = thin_border
            cur += 1

            # Data
            for r_off, (_, row) in enumerate(df.iterrows()):
                fill = alt_fill if r_off % 2 else wht_fill
                for c_idx, val in enumerate(row, 1):
                    dc = ws.cell(cur, c_idx, clean_cell(val))
                    dc.fill = fill
                    dc.font = data_font
                    dc.border = thin_border
                    dc.alignment = Alignment(
                        horizontal="left" if c_idx == 1 else "center",
                        vertical="center",
                        wrap_text=True,
                    )
                    # Attempt to format as currency if it contains $
                    if "$" in str(val):
                        try:
                            # Remove $ and words, keep number
                            num_str = re.sub(
                                r"[^\d\.]",
                                "",
                                str(val).split()[0] if " " in str(val) else str(val),
                            )
                            if num_str:
                                dc.number_format = '"$"#,##0.00'
                        except Exception:
                            pass
                cur += 1

            cur += 3  # more spacer

        # Column widths with better sizing
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            max_len = max((len(str(c.value or "")) for c in col_cells), default=12)
            ws.column_dimensions[col_letter].width = min(max(max_len * 1.2, 15), 70)

        ws.freeze_panes = "A4"  # Freeze after company header

    wb.save(out_path)


# ── Main ───────────────────────────────────────────────────────────────────────


def main():
    print("Mining Capex Table Extractor")
    print("=" * 55)
    print("Strategy: TABLE -> Excel  |  No table -> .txt file\n")

    table_data: dict[str, list[dict]] = {}  # companies with real tables
    text_data: dict[str, list[dict]] = {}  # companies with text fallback

    for company, pdf_path in TARGETS.items():
        print(f"[{company}]")

        # Step 1 — try tables
        tables = try_extract_tables(pdf_path)

        if tables:
            print(f"  Table extraction: {len(tables)} capex table(s) found")
            for t in tables:
                print(f"    pg{t['page']} | {t['df'].shape[0]}r x {t['df'].shape[1]}c")
            table_data[company] = tables
        else:
            # Step 2 — fallback to text file
            print("  No tables found — extracting text with $ amounts -> .txt")
            fallback = fallback_text_extract(pdf_path)
            print(f"  Found: {len(fallback)} capex paragraph(s)")
            text_data[company] = fallback

    # ── Save Excel (table companies only) ─────────────────────────────────────
    if table_data:
        out_excel = Path("capex_tables.xlsx")
        write_excel(table_data, out_excel)
        print(f"\nExcel saved -> {out_excel}")

    # ── Save .txt files (no-table companies) ──────────────────────────────────
    for company, entries in text_data.items():
        safe_name = company.replace(" ", "_")
        txt_path = Path(f"{safe_name}_capex.txt")
        with txt_path.open("w", encoding="utf-8") as f:
            f.write(f"Capital Expenditure Data — {company}\n")
            f.write("=" * 60 + "\n")
            f.write(
                "(No structured table found in PDF. Extracted paragraphs with $ amounts.)\n\n"
            )
            for entry in entries:
                f.write(f"--- Page {entry['page']} ---\n")
                f.write(entry["text"] + "\n\n")
        print(f"Text saved  -> {txt_path}")

    # ── Summary ────────────────────────────────────────────────────────────────
    print("\n" + "=" * 55)
    print(f"{'Company':<22} {'Output':<30} {'Entries':>5}")
    print("-" * 58)
    for company, entries in table_data.items():
        print(f"{company:<22} {'capex_tables.xlsx':<30} {len(entries):>5}")
    for company, entries in text_data.items():
        fname = f"{company.replace(' ','_')}_capex.txt"
        print(f"{company:<22} {fname:<30} {len(entries):>5}")


if __name__ == "__main__":
    main()
